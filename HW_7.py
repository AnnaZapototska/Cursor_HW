#task 1.1
with open("task1.txt", "r") as file:
    content = file.readlines()
    dict = {}
    for i in range(0, len(content), 2):
        content[i] = content[i].replace("\n", "")
        content[i + 1] = content[i + 1].replace("\n", "")
        dict.update({content[i]: content[i + 1]})
    print(dict)

#task 1.2
with open('task1.2.txt', 'w') as file:
        for item in dict.values():
            file.write(f'{item} ')

#file = open('task1.2.txt', 'r')
#b = file.read()
#print(b)

#task 2
import pickle
with open("task2", "rb") as file:
    list = pickle.load(file)
    average = sum(list)/len(list)
    print(average)

#task 3
import openpyxl


class CreateWorkbook:
    def __init__(self, file_name):
        self.file_obj = openpyxl.load_workbook(file_name)

    def __enter__(self):
        return self.file_obj

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.file_obj.close()

with CreateWorkbook('Task3.xlsx') as file:
    print(file.sheetnames)
    sheet = file.active
    print(sheet['A3'].value, sheet['B3'].value)
    sheet['A4'] = 25
    print(sheet['A4'].value)
    file.save('Task3_update.xlsx')

